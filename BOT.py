import discord
import openpyxl
import os
import re
from datetime import datetime
from dotenv import load_dotenv

load_dotenv()

TOKEN = os.getenv("DISCORD_TOKEN")
CHANNEL_NAME = "🤝自己紹介"        # 監視するチャンネル名（変えてOK）
EXCEL_FILE = "自己紹介.xlsx"

# ===== Excelファイルの初期化 =====
def init_excel():
    if not os.path.exists(EXCEL_FILE):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "🤝自己紹介"
        ws.append([
            "投稿日時",
            "Discordユーザー名",
            "本名",
            "学籍番号",
            "SNSアカウント",
            "好きなゲーム",
            "じょぎでやりたいこと",
            "ひとこと",
        ])
        # 列幅の調整
        col_widths = [18, 20, 15, 15, 25, 20, 30, 30]
        for i, width in enumerate(col_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
        wb.save(EXCEL_FILE)
        print(f"[初期化] {EXCEL_FILE} を作成しました")

# ===== 自己紹介テキストのパース =====
def parse_intro(text: str) -> dict:
    fields = {
        "本名": "",
        "学籍番号": "",
        "SNSアカウント": "",
        "好きなゲーム": "",
        "じょぎでやりたいこと": "",
        "ひとこと": "",
    }

    # 各項目のキーワードと次の項目のキーワードで範囲を切り出す
    patterns = {
        "本名":             r"本名[：:]\s*(.+?)(?=○|$)",
        "学籍番号":         r"学籍番号[：:]\s*(.+?)(?=○|$)",
        "SNSアカウント":    r"SNSアカウント[：:]\s*(.+?)(?=○|$)",
        "好きなゲーム":     r"好きなゲーム[：:]\s*(.+?)(?=○|$)",
        "じょぎでやりたいこと": r"じょぎでやりたいこと[：:]\s*(.+?)(?=○|$)",
        "ひとこと":         r"ひとこと[：:]\s*(.+?)(?=○|$)",
    }

    for key, pattern in patterns.items():
        match = re.search(pattern, text, re.DOTALL)
        if match:
            value = match.group(1).strip()
            # 改行を空白に整形
            value = re.sub(r"\s+", " ", value)
            fields[key] = value

    return fields

# ===== Excelへの書き込み =====
def write_to_excel(discord_username: str, fields: dict, timestamp: datetime):
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    ws.append([
        timestamp.strftime("%Y-%m-%d %H:%M:%S"),
        discord_username,
        fields["本名"],
        fields["学籍番号"],
        fields["SNSアカウント"],
        fields["好きなゲーム"],
        fields["じょぎでやりたいこと"],
        fields["ひとこと"],
    ])
    wb.save(EXCEL_FILE)
    print(f"[記録] {discord_username} の自己紹介を書き込みました")

# ===== Discord Bot =====
intents = discord.Intents.default()
intents.message_content = True  # メッセージ内容の読み取りに必要

client = discord.Client(intents=intents)

@client.event
async def on_ready():
    init_excel()
    print(f"[起動] {client.user} としてログイン完了")
    print(f"[監視] チャンネル名「{CHANNEL_NAME}」を監視中...")

@client.event
async def on_message(message: discord.Message):
    # Bot自身の投稿は無視
    if message.author.bot:
        return

    # 対象チャンネルかチェック
    if message.channel.name != CHANNEL_NAME:
        return

    content = message.content

    # 本名か学籍番号が含まれているか最低限チェック
    if "本名" not in content and "学籍番号" not in content:
        return

    fields = parse_intro(content)

    # 本名・学籍番号が両方空なら記録しない
    if not fields["本名"] and not fields["学籍番号"]:
        print(f"[スキップ] {message.author} の投稿：本名・学籍番号が取得できませんでした")
        return

    write_to_excel(
        discord_username=str(message.author),
        fields=fields,
        timestamp=message.created_at.replace(tzinfo=None),
    )

    # 任意：Botがリアクションで完了を通知
    await message.add_reaction("✅")

client.run(TOKEN)